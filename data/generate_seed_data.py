#!/usr/bin/env python3
"""
Aurelius Wealth Management — Seed Data Generator (v2 — fully relational)
All categorical fields in entity tables are FK integers pointing to a lookup table.
Run:  python generate_seed_data.py
"""
import os, datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE     = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(HERE, "aurelius_seed_data.xlsx")
D        = datetime.date

# ── Styling ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1C2130")
HDR_FONT  = Font(bold=True, color="C49A38", name="Calibri", size=10)
BODY_FONT = Font(name="Calibri", size=10)
LKP_FILL  = PatternFill("solid", fgColor="0D3349")   # teal-navy for lookup sheets
LKP_FONT  = Font(bold=True, color="7FB8D4", name="Calibri", size=10)

def _style(ws, is_lookup=False):
    fill = LKP_FILL if is_lookup else HDR_FILL
    font = LKP_FONT if is_lookup else HDR_FONT
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2,10),55)
    ws.freeze_panes = "A2"
    if ws.max_row > 1: ws.auto_filter.ref = ws.dimensions

def sheet(wb, name, headers, rows, lookup=False):
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for r in rows: ws.append(r)
    _style(ws, lookup)
    return ws


# =============================================================================
# PART 1 — LOOKUP TABLES
# =============================================================================

lkp_title = [
    (1,"Mr"),(2,"Mrs"),(3,"Ms"),(4,"Miss"),(5,"Dr"),
    (6,"Prof"),(7,"Rev"),(8,"Sir"),(9,"Lady"),(10,"Lord"),
]
lkp_gender = [(1,"Male"),(2,"Female"),(3,"Non-binary"),(4,"Prefer not to say")]

lkp_marital = [
    (1,"Single"),(2,"Married"),(3,"Divorced"),(4,"Widowed"),
    (5,"Separated"),(6,"Civil Partnership"),(7,"Cohabiting"),
]
lkp_nationality = [
    (1,"British","United Kingdom"),(2,"Irish","Ireland"),(3,"French","France"),
    (4,"German","Germany"),(5,"American","United States"),(6,"Indian","India"),
    (7,"Pakistani","Pakistan"),(8,"Australian","Australia"),(9,"Canadian","Canada"),
    (10,"South African","South Africa"),(11,"Italian","Italy"),(12,"Spanish","Spain"),
    (13,"Polish","Poland"),(14,"Dutch","Netherlands"),(15,"Swedish","Sweden"),
    (16,"Norwegian","Norway"),(17,"Danish","Denmark"),(18,"Swiss","Switzerland"),
    (19,"Belgian","Belgium"),(20,"Portuguese","Portugal"),
]
lkp_country = [
    (1,"United Kingdom","GB","GBR"),(2,"Ireland","IE","IRL"),
    (3,"France","FR","FRA"),(4,"Germany","DE","DEU"),
    (5,"United States","US","USA"),(6,"India","IN","IND"),
    (7,"Australia","AU","AUS"),(8,"Canada","CA","CAN"),
    (9,"Spain","ES","ESP"),(10,"Italy","IT","ITA"),
    (11,"Portugal","PT","PRT"),(12,"Switzerland","CH","CHE"),
    (13,"Sweden","SE","SWE"),(14,"Norway","NO","NOR"),
    (15,"Denmark","DK","DNK"),(16,"Netherlands","NL","NLD"),
    (17,"Belgium","BE","BEL"),(18,"Poland","PL","POL"),
    (19,"South Africa","ZA","ZAF"),(20,"New Zealand","NZ","NZL"),
]
lkp_branch = [
    (1,"London City","London","14 St James's Square","SW1Y 4LB"),
    (2,"Edinburgh","Edinburgh","3 Charlotte Square","EH2 4DR"),
    (3,"Bristol","Bristol","8 Whiteladies Road","BS8 1PD"),
    (4,"Manchester","Manchester","24 Deansgate","M3 4LY"),
]
lkp_adviser_role = [
    (1,"Associate Wealth Manager",1),
    (2,"Wealth Manager",2),
    (3,"Senior Wealth Manager",3),
    (4,"Principal Wealth Manager",4),
    (5,"Partner",5),
]
lkp_tier = [
    (1,"Bronze",0,"New clients or portfolios under £50,000"),
    (2,"Silver",50000,"Portfolios £50,000 – £250,000"),
    (3,"Gold",250000,"Portfolios £250,000 – £750,000"),
    (4,"Platinum",750000,"Portfolios over £750,000 — dedicated adviser and priority service"),
]
lkp_kyc_status = [
    (1,"Verified","Identity verified and documentation current"),
    (2,"Pending","Verification in progress"),
    (3,"Expired","Documentation expired — renewal required"),
    (4,"Rejected","Verification failed — re-submission required"),
]
lkp_kyc_doc = [
    (1,"Passport"),(2,"Driving Licence"),(3,"National Identity Card"),
    (4,"Biometric Residence Permit"),(5,"HM Forces ID Card"),
]
lkp_account_type = [
    (1,"ISA","Individual Savings Account",20000,"Tax-free up to £20,000 per tax year"),
    (2,"GIA","General Investment Account",0,"No tax wrapper — gains and income are taxable"),
    (3,"SIPP","Self-Invested Personal Pension",60000,"Tax relief on contributions. Annual allowance £60,000"),
    (4,"JISA","Junior Individual Savings Account",9000,"Tax-free for under-18s. Up to £9,000 per tax year"),
    (5,"LISA","Lifetime Individual Savings Account",4000,"25% government bonus on contributions up to £4,000"),
]
lkp_account_status = [
    (1,"Active"),(2,"Dormant"),(3,"Closed"),(4,"Suspended"),(5,"Pending Closure"),
]
lkp_fund_type = [
    (1,"ETF","Exchange Traded Fund — passively or actively managed, listed on exchange"),
    (2,"MutualFund","Pooled investment fund — priced once daily, bought/sold via fund manager"),
    (3,"Share","Individual company equity listed on a stock exchange"),
    (4,"Bond","Fixed income debt security issued by government or corporation"),
    (5,"Cash","Cash or near-cash equivalent held on deposit"),
]
lkp_asset_class = [
    (1,"Global Equity"),(2,"UK Equity"),(3,"US Equity"),(4,"European Equity"),
    (5,"Emerging Markets Equity"),(6,"Asia Pacific Equity"),(7,"Japan Equity"),
    (8,"Mixed Asset"),(9,"Fixed Income — Government"),(10,"Fixed Income — Corporate"),
    (11,"Property"),(12,"Infrastructure"),(13,"Commodities"),(14,"Cash"),(15,"Alternative"),
]
lkp_currency = [
    (1,"GBP","British Pound Sterling","£"),(2,"USD","United States Dollar","$"),
    (3,"EUR","Euro","€"),(4,"JPY","Japanese Yen","¥"),(5,"CHF","Swiss Franc","Fr"),
    (6,"AUD","Australian Dollar","A$"),(7,"CAD","Canadian Dollar","C$"),
    (8,"SEK","Swedish Krona","kr"),(9,"NOK","Norwegian Krone","kr"),
    (10,"HKD","Hong Kong Dollar","HK$"),
]
lkp_txn_type = [
    (1,"Buy",False,"Purchase of units/shares — debits cash"),
    (2,"Sell",True,"Sale of units/shares — credits cash"),
    (3,"Dividend",True,"Income distribution from a holding"),
    (4,"Deposit",True,"Cash deposited into account"),
    (5,"Withdrawal",False,"Cash withdrawn from account"),
    (6,"Transfer In",True,"Assets transferred in from another provider"),
    (7,"Transfer Out",False,"Assets transferred out to another provider"),
    (8,"Management Fee",False,"Platform or adviser fee deducted"),
    (9,"Interest",True,"Interest credited on cash balance"),
]
lkp_txn_status = [
    (1,"Settled"),(2,"Pending"),(3,"Cancelled"),(4,"Failed"),(5,"In Progress"),
]
lkp_doc_cat = [
    (1,"Statement","Periodic portfolio or account valuation statement"),
    (2,"Tax","Tax certificates, CGT summaries, dividend vouchers"),
    (3,"Correspondence","Letters and secure messages"),
    (4,"Report","Quarterly and annual investment reports, suitability reports"),
    (5,"Form","Account opening forms, instructions, agreements"),
]
lkp_addr_type = [
    (1,"Registered","Primary legal address for regulatory purposes"),
    (2,"Correspondence","Preferred mailing address if different from registered"),
    (3,"Previous","Former address retained for audit trail"),
    (4,"Overseas","Non-UK address"),
]
lkp_contact_method = [
    (1,"Email"),(2,"Phone"),(3,"Post"),(4,"SMS"),
]
lkp_notif_channel = [
    (1,"Email"),(2,"SMS"),(3,"InApp"),
]
lkp_notif_category = [
    (1,"Statements","Portfolio and account statements"),
    (2,"Trades","Trade confirmations and dealing instructions"),
    (3,"Markets","Market news and investment commentary"),
    (4,"Reviews","Upcoming and completed annual reviews"),
    (5,"System","Security alerts, login activity, platform notices"),
]
lkp_life_stage = [
    (1,"Starting out","Student or early career — building financial foundations"),
    (2,"Early career","First decade of working life — beginning to invest"),
    (3,"Building wealth","Mid-career with growing assets and disposable income"),
    (4,"Peak earning","Prime earning years — maximising accumulation"),
    (5,"Pre-retirement","Within 10 years of retirement — capital preservation focus"),
    (6,"Retirement","Drawing on accumulated wealth — income generation priority"),
]
lkp_risk = [
    (1,"Preserve","Capital protection above all — minimal risk, accepts lower returns"),
    (2,"Cautious","Modest growth with capital preservation — low volatility portfolio"),
    (3,"Balanced","Growth and stability balanced — accepts moderate market swings"),
    (4,"Growth","Long-term capital growth — accepts significant short-term volatility"),
    (5,"Aggressive","Maximum long-term return — fully accepts high volatility and drawdown"),
]
lkp_priority = [
    (1,"Capital preservation & security"),
    (2,"Long-term wealth accumulation"),
    (3,"Financial independence & freedom"),
    (4,"Values-aligned & ethical investing"),
    (5,"Income generation & yield"),
]
lkp_capital_range = [
    (1,"No capital",     0,      0,      "No investable capital at this time"),
    (2,"Under £500",     0,      500,    "Under £500"),
    (3,"£500 – £2,000",  500,    2000,   "£500 to £2,000"),
    (4,"£2,000 – £10,000",2000,  10000,  "£2,000 to £10,000"),
    (5,"£10,000 – £50,000",10000,50000,  "£10,000 to £50,000"),
    (6,"Over £50,000",   50000,  999999, "Over £50,000"),
]
lkp_element = [
    (1,"Water","#4a8fe8","Flow and depth — patient, analytical, long-term perspective"),
    (2,"Wood", "#3daa6a","Growth and resilience — structured, progressive, disciplined"),
    (3,"Fire", "#e06a45","Energy and boldness — decisive, opportunistic, high conviction"),
    (4,"Earth","#c49a38","Stability and nurture — conservative, balanced, preservation-focused"),
    (5,"Metal","#6e8aaa","Precision and clarity — systematic, data-driven, risk-aware"),
]
lkp_theme = [(1,"Dark"),(2,"Light"),(3,"System")]
lkp_tax_year = [
    (1,"2022/23",D(2022,4,6),D(2023,4,5),20000,40000,6000),
    (2,"2023/24",D(2023,4,6),D(2024,4,5),20000,60000,6000),
    (3,"2024/25",D(2024,4,6),D(2025,4,5),20000,60000,3000),
    (4,"2025/26",D(2025,4,6),D(2026,4,5),20000,60000,3000),
]


# =============================================================================
# PART 2 — ENTITY TABLES  (FK integer IDs throughout)
# =============================================================================

# ── Adviser ───────────────────────────────────────────────────────────────────
# RoleID: 2=Wealth Manager  3=Senior Wealth Manager
# BranchID: 1=London  2=Edinburgh  3=Bristol  4=Manchester
adviser_rows = [
    (1,"James",  "Whitmore",      "j.whitmore@aurelius.co.uk",      3,"+44 20 7946 0100", 1,True),
    (2,"Sarah",  "Hollingsworth","s.hollingsworth@aurelius.co.uk",  3,"+44 131 946 0200", 2,True),
    (3,"Robert", "Chambers",     "r.chambers@aurelius.co.uk",       2,"+44 117 946 0300", 3,True),
    (4,"Eleanor","Forsyth",      "e.forsyth@aurelius.co.uk",        3,"+44 20 7946 0400", 1,True),
    (5,"Andrew", "MacPhail",     "a.macphail@aurelius.co.uk",       2,"+44 161 946 0500", 4,True),
]
adviser_hdr = ["AdviserID","FirstName","LastName","Email","RoleID","Phone","BranchID","IsActive"]

# ── Client ────────────────────────────────────────────────────────────────────
# TitleID: 1=Mr 2=Mrs 3=Ms 4=Miss 5=Dr
# GenderID: 1=Male 2=Female
# NationalityID: 1=British 6=Indian
# MaritalStatusID: 1=Single 2=Married 3=Divorced 4=Widowed
# TaxResidencyCountryID: 1=United Kingdom
# TierID: 1=Bronze 2=Silver 3=Gold 4=Platinum
# KYCStatusID: 1=Verified
# KYCDocTypeID: 1=Passport 2=Driving Licence
client_rows = [
    (1, "AW-20240-UK",3, 1,1,"Oliver",   "Pemberton",  "Oliver",  D(2002,3,14),1,1,1, "PO123456A",1,"2847391650",False,False,D(2024,1,15),1,D(2026,7,15), 1,1,D(2031,3,14),D(2024,1,15)),
    (2, "AW-20192-UK",5, 3,2,"Priya",    "Sharma",     "Priya",   D(1995,7,22),2,1,1, "PS234567B",1,"3916284750",False,False,D(2019,2,10),2,D(2026,8,10), 1,1,D(2030,7,22),D(2019,2,10)),
    (3, "AW-20174-UK",5, 1,1,"Thomas",   "Blackwood",  "Thomas",  D(1988,1, 9),1,1,2, "TB345678C",1,"4827391650",False,False,D(2017,4,20),2,D(2026,10,20),1,1,D(2033,1, 9),D(2023,4,20)),
    (4, "AW-20156-UK",3, 2,2,"Charlotte","Davies",     "Charlotte",D(1983,5,16),2,1,2, "CD456789D",1,"5918273460",False,False,D(2015,6,12),3,D(2026,12,12),1,1,D(2028,5,16),D(2025,6,12)),
    (5, "AW-20123-UK",5, 1,1,"James",    "Holroyd",    "James",   D(1978,11,3),1,1,2, "JH567890A",1,"6284739150",False,False,D(2012,3, 8),3,D(2026,9, 8), 1,2,D(2029,11,3),D(2024,3, 8)),
    (6, "AW-20182-UK",1, 1,1,"Sumant",   "Kumar",      "Sumant",  D(1978,7,14),1,6,2, "AB123456C",1,"1234567890",False,False,D(2018,3, 1),4,D(2026,9, 1), 1,1,D(2031,11,24),D(2024,1,18)),
    (7, "AW-20098-UK",2, 3,2,"Fiona",    "Mackintosh", "Fiona",   D(1971,4,28),2,1,3, "FM678901B",1,"7391826450",False,False,D(2009,9,14),4,D(2026,9,14), 1,1,D(2031,4,28),D(2024,9,14)),
    (8, "AW-20051-UK",3, 1,1,"Richard",  "Ashworth",   "Richard", D(1963,9, 7),1,1,2, "RA789012C",1,"8274918360",False,False,D(2005,1,22),4,D(2026,7,22), 1,1,D(2028,9, 7),D(2025,1,22)),
    (9, "AW-19978-UK",4, 2,2,"Margaret", "Thornton",   "Margaret",D(1955,2,19),2,1,4, "MT890123D",1,"9163728450",False,False,D(1997,8, 5),4,D(2026,11,5), 1,1,D(2030,2,19),D(2025,8, 5)),
    (10,"AW-20073-UK",4, 1,1,"David",    "Winterbottom","David",  D(1968,6,30),1,1,2, "DW901234A",1,"1827364950",False,False,D(2007,3,17),4,D(2026,9,17), 1,1,D(2033,6,30),D(2024,3,17)),
]
client_hdr = [
    "ClientID","ClientNumber","AdviserID",
    "TitleID","GenderID","FirstName","LastName","PreferredName","DateOfBirth",
    "NationalityID","TaxResidencyCountryID","MaritalStatusID",
    "NationalInsuranceNumber","TaxResidencyCountryID2","UniqueTaxpayerReference",
    "FATCADeclaration","USPersonDeclaration",
    "ClientSince","TierID","NextReviewDate",
    "KYCStatusID","KYCDocTypeID","KYCDocumentExpiry","KYCVerifiedDate",
]
# Fix duplicate column name
client_hdr[13] = "TaxResidencyCountryID"
client_hdr[12] = "NationalInsuranceNumber"

# rebuild cleanly
client_hdr = [
    "ClientID","ClientNumber","AdviserID",
    "TitleID","GenderID","FirstName","LastName","PreferredName","DateOfBirth",
    "NationalityID","TaxResidencyCountryID","MaritalStatusID",
    "NationalInsuranceNumber","UniqueTaxpayerReference",
    "FATCADeclaration","USPersonDeclaration",
    "ClientSince","TierID","NextReviewDate",
    "KYCStatusID","KYCDocTypeID","KYCDocumentExpiry","KYCVerifiedDate",
]
# Re-align rows (drop the duplicate col that was in previous draft)
client_rows = [
    (1, "AW-20240-UK",3, 1,1,"Oliver",   "Pemberton",  "Oliver",  D(2002,3,14),1,1,1,"PO123456A","2847391650",False,False,D(2024,1,15),1,D(2026,7,15), 1,1,D(2031,3,14), D(2024,1,15)),
    (2, "AW-20192-UK",5, 3,2,"Priya",    "Sharma",     "Priya",   D(1995,7,22),1,1,1,"PS234567B","3916284750",False,False,D(2019,2,10),2,D(2026,8,10), 1,1,D(2030,7,22), D(2019,2,10)),
    (3, "AW-20174-UK",5, 1,1,"Thomas",   "Blackwood",  "Thomas",  D(1988,1, 9),1,1,2,"TB345678C","4827391650",False,False,D(2017,4,20),2,D(2026,10,20),1,1,D(2033,1, 9),  D(2023,4,20)),
    (4, "AW-20156-UK",3, 2,2,"Charlotte","Davies",     "Charlotte",D(1983,5,16),1,1,2,"CD456789D","5918273460",False,False,D(2015,6,12),3,D(2026,12,12),1,1,D(2028,5,16), D(2025,6,12)),
    (5, "AW-20123-UK",5, 1,1,"James",    "Holroyd",    "James",   D(1978,11,3),1,1,2,"JH567890A","6284739150",False,False,D(2012,3, 8),3,D(2026,9, 8), 1,2,D(2029,11,3), D(2024,3, 8)),
    (6, "AW-20182-UK",1, 1,1,"Sumant",   "Kumar",      "Sumant",  D(1978,7,14),6,1,2,"AB123456C","1234567890",False,False,D(2018,3, 1),4,D(2026,9, 1), 1,1,D(2031,11,24),D(2024,1,18)),
    (7, "AW-20098-UK",2, 3,2,"Fiona",    "Mackintosh", "Fiona",   D(1971,4,28),1,1,3,"FM678901B","7391826450",False,False,D(2009,9,14),4,D(2026,9,14), 1,1,D(2031,4,28), D(2024,9,14)),
    (8, "AW-20051-UK",3, 1,1,"Richard",  "Ashworth",   "Richard", D(1963,9, 7),1,1,2,"RA789012C","8274918360",False,False,D(2005,1,22),4,D(2026,7,22), 1,1,D(2028,9, 7),  D(2025,1,22)),
    (9, "AW-19978-UK",4, 2,2,"Margaret", "Thornton",   "Margaret",D(1955,2,19),1,1,4,"MT890123D","9163728450",False,False,D(1997,8, 5),4,D(2026,11,5), 1,1,D(2030,2,19), D(2025,8, 5)),
    (10,"AW-20073-UK",4, 1,1,"David",    "Winterbottom","David",  D(1968,6,30),1,1,2,"DW901234A","1827364950",False,False,D(2007,3,17),4,D(2026,9,17), 1,1,D(2033,6,30), D(2024,3,17)),
]

# ── ClientContact ─────────────────────────────────────────────────────────────
# PreferredContactMethodID: 1=Email 2=Phone 3=Post 4=SMS
contact_rows = [
    (1, 1, "oliver.pemberton@gmail.com",        True, "",                                 "+44 7700 900 101","+44 117 946 0101","",                  1),
    (2, 2, "priya.sharma@outlook.com",           True, "p.sharma@sharmatech.co.uk",       "+44 7700 900 102","",                 "+44 161 946 0202", 1),
    (3, 3, "thomas.blackwood@gmail.com",         True, "t.blackwood@bwarchitects.co.uk",  "+44 7700 900 103","+44 113 946 0303","+44 113 946 0304",  1),
    (4, 4, "charlotte.davies@gmail.com",         True, "charlotte@daviesconsulting.co.uk","+44 7700 900 104","+44 29 946 0404", "+44 29 946 0405",   1),
    (5, 5, "james.holroyd@gmail.com",            True, "j.holroyd@holroydproperties.co.uk","+44 7700 900 105","+44 1423 946 505","+44 1423 946 506", 1),
    (6, 6, "sumant.kumar@gmail.com",             True, "s.kumar@kumar-ventures.co.uk",    "+44 7700 900 142","+44 20 7946 0352","",                  1),
    (7, 7, "fiona.mackintosh@gmail.com",         True, "f.mackintosh@mackcapital.co.uk",  "+44 7700 900 107","+44 131 946 0707","",                  1),
    (8, 8, "richard.ashworth@gmail.com",         True, "richard@ashworthfamily.co.uk",    "+44 7700 900 108","+44 1225 946 808","",                  2),
    (9, 9, "margaret.thornton@btinternet.com",   True, "",                                 "+44 7700 900 109","+44 1962 946 909","",                  3),
    (10,10,"david.winterbottom@gmail.com",       True, "d.winterbottom@wintersgroup.co.uk","+44 7700 900 110","+44 1483 946 010","+44 1483 946 011", 1),
]
contact_hdr = ["ContactID","ClientID","EmailPrimary","EmailPrimaryVerified",
               "EmailSecondary","PhoneMobile","PhoneHome","PhoneWork",
               "PreferredContactMethodID"]

# ── ClientAddress ─────────────────────────────────────────────────────────────
# AddressTypeID: 1=Registered 2=Correspondence
# CountryID: 1=United Kingdom
address_rows = [
    (1,  1, 1,"8 Whiteladies Road",       "",                   "Bristol",    "Avon",              "BS8 1PD", 1,True),
    (2,  2, 1,"24 Deansgate",             "Flat 12",            "Manchester", "Greater Manchester","M3 4LY",  1,True),
    (3,  3, 1,"15 The Calls",             "",                   "Leeds",      "West Yorkshire",    "LS2 7EY", 1,True),
    (4,  4, 1,"6 Cathedral Road",         "",                   "Cardiff",    "South Glamorgan",   "CF11 9LJ",1,True),
    (5,  4, 2,"c/o Davies Consulting",    "12 St Mary Street",  "Cardiff",    "South Glamorgan",   "CF10 1AT",1,True),
    (6,  5, 1,"12 Parliament Street",     "",                   "Harrogate",  "North Yorkshire",   "HG1 2QU", 1,True),
    (7,  6, 1,"14 Cavendish Square",      "",                   "London",     "",                  "W1G 0PH", 1,True),
    (8,  7, 1,"3 Charlotte Square",       "Flat 4",             "Edinburgh",  "Midlothian",        "EH2 4DR", 1,True),
    (9,  8, 1,"22 Great Pulteney Street", "",                   "Bath",       "Somerset",          "BA2 4BR", 1,True),
    (10, 9, 1,"8 The Square",             "",                   "Winchester", "Hampshire",         "SO23 9ES",1,True),
    (11,10, 1,"45 High Street",           "",                   "Guildford",  "Surrey",            "GU1 3DL", 1,True),
    (12,10, 2,"Winters Group Ltd",        "2 Chertsey Street",  "Guildford",  "Surrey",            "GU1 4HD", 1,True),
]
address_hdr = ["AddressID","ClientID","AddressTypeID","AddressLine1","AddressLine2",
               "City","County","Postcode","CountryID","IsActive"]

# ── ClientPreferences ─────────────────────────────────────────────────────────
# AppThemeID: 1=Dark 2=Light 3=System
pref_rows = [
    (1, 1, True, "English","Europe/London",False,False,True, 1),
    (2, 2, True, "English","Europe/London",True, False,True, 2),
    (3, 3, True, "English","Europe/London",False,False,True, 1),
    (4, 4, True, "English","Europe/London",False,False,True, 2),
    (5, 5, False,"English","Europe/London",False,False,False,2),
    (6, 6, True, "English","Europe/London",False,False,True, 1),
    (7, 7, True, "English","Europe/London",True, False,True, 1),
    (8, 8, False,"English","Europe/London",False,False,False,2),
    (9, 9, False,"English","Europe/London",False,False,False,2),
    (10,10,True, "English","Europe/London",False,False,True, 1),
]
pref_hdr = ["PreferenceID","ClientID","Paperless","Language","Timezone",
            "MarketingConsent","ThirdPartyConsent","AnalyticsConsent","AppThemeID"]

# ── NotificationPreferences ───────────────────────────────────────────────────
# ChannelID: 1=Email 2=SMS 3=InApp
# One row per client per channel; boolean columns per category
notif_cfg = {
    #     Email(S,T,M,R,Sy)         SMS(S,T,M,R,Sy)          InApp(S,T,M,R,Sy)
    1: [(True,True,False,True,True), (False,True,False,False,False),(True,True,True,True,True)],
    2: [(True,True,True,True,True),  (False,True,False,False,True), (True,True,True,True,True)],
    3: [(True,True,False,True,True), (False,True,False,False,False),(True,True,True,True,True)],
    4: [(True,True,False,True,True), (False,False,False,False,False),(True,True,True,True,True)],
    5: [(True,True,True,True,True),  (True,True,False,True,False),  (True,True,True,True,True)],
    6: [(True,True,False,True,True), (False,True,False,False,False),(True,True,True,True,True)],
    7: [(True,True,True,True,True),  (False,False,False,True,False),(True,True,True,True,True)],
    8: [(True,False,False,True,True),(False,False,False,False,False),(True,True,False,True,True)],
    9: [(True,False,False,True,True),(False,False,False,False,False),(False,False,False,True,False)],
    10:[(True,True,True,True,True),  (False,True,False,True,False), (True,True,True,True,True)],
}
notif_rows, npid = [], 1
for cid, (ecfg, scfg, icfg) in notif_cfg.items():
    notif_rows.append((npid,cid,1,*ecfg)); npid+=1
    notif_rows.append((npid,cid,2,*scfg)); npid+=1
    notif_rows.append((npid,cid,3,*icfg)); npid+=1
notif_hdr = ["NotifPrefID","ClientID","ChannelID",
             "Statements","Trades","Markets","Reviews","System"]

# ── InvestmentProfile ─────────────────────────────────────────────────────────
# LifeStageID 1-6, RiskToleranceID 1-5
# WealthPriorityID_1-5 → FK to lkp_priority (1-5)
# CapitalRangeID: 5=£10k-50k  6=£50k+
# DominantElementID: 1=Water 2=Wood 3=Fire 4=Earth 5=Metal
profile_rows = [
    (1, 1, D(2024,1,20),1,4,2,4,3,1,5,6,2,60,35,68,42,30,25,True),
    (2, 2, D(2019,2,15),2,3,3,2,4,5,1,6,3,40,38,62,35,25,True),  # note: only 5 priority IDs
    (3, 3, D(2023,4,25),3,4,2,1,3,5,4,6,1,72,45,55,40,38,True),
    (4, 4, D(2025,6,15),3,3,1,5,2,3,4,6,4,45,52,38,75,40,True),
    (5, 5, D(2024,3,10),4,3,2,5,1,3,4,6,5,42,38,45,55,80,True),
    (6, 6, D(2024,1,22),3,3,2,1,3,5,4,6,1,68,42,38,48,52,True),
    (7, 7, D(2024,9,18),5,2,1,5,2,3,4,6,4,55,48,32,82,43,True),
    (8, 8, D(2025,1,25),5,2,5,1,2,3,4,6,5,38,42,28,62,85,True),
    (9, 9, D(2025,8,10),6,1,1,5,2,4,3,6,4,28,32,22,88,55,True),
    (10,10,D(2024,3,20),4,3,2,1,5,3,4,6,1,78,52,48,62,60,True),
]
# cols: ProfileID,ClientID,AssessmentDate, LifeStageID,RiskToleranceID,
#       WealthPriorityID1-5, CapitalRangeID, DominantElementID,
#       ScoreWater,Wood,Fire,Earth,Metal, IsActive
profile_hdr = [
    "ProfileID","ClientID","AssessmentDate",
    "LifeStageID","RiskToleranceID",
    "WealthPriorityID1","WealthPriorityID2","WealthPriorityID3",
    "WealthPriorityID4","WealthPriorityID5",
    "CapitalRangeID","DominantElementID",
    "ElementScoreWater","ElementScoreWood","ElementScoreFire",
    "ElementScoreEarth","ElementScoreMetal","IsActive",
]

# ── Account ───────────────────────────────────────────────────────────────────
# AccountTypeID: 1=ISA 2=GIA 3=SIPP  AccountStatusID: 1=Active  TaxYearID: 4=2025/26
account_rows = [
    (1, 1, "ISA-20240-001",1,D(2024,1,15),1,  2500.00,  10000.00,12500.00,4),
    (2, 2, "ISA-20192-001",1,D(2019,2,10),1,  1200.00,  30800.00,20000.00,4),
    (3, 2, "SIP-20192-001",3,D(2019,2,10),1,   800.00,  12400.00,    0.00,4),
    (4, 3, "ISA-20174-001",1,D(2017,4,20),1,  3400.00,  36600.00,20000.00,4),
    (5, 3, "SIP-20174-001",3,D(2017,4,20),1,  4200.00,  67800.00,    0.00,4),
    (6, 3, "GIA-20174-001",2,D(2020,6,10),1,  1800.00,  14600.00,    0.00,4),
    (7, 4, "ISA-20156-001",1,D(2015,6,12),1,  5000.00,  80000.00,15000.00,4),
    (8, 4, "SIP-20156-001",3,D(2015,6,12),1,  8000.00, 152000.00,    0.00,4),
    (9, 4, "GIA-20156-001",2,D(2018,9, 5),1,  2000.00,  37000.00,    0.00,4),
    (10,5, "ISA-20123-001",1,D(2012,3, 8),1,  4000.00,  88000.00,20000.00,4),
    (11,5, "SIP-20123-001",3,D(2012,3, 8),1, 12000.00, 321000.00,    0.00,4),
    (12,6, "SIP-20182-001",3,D(2018,3, 1),1, 31800.00, 810540.00,    0.00,4),
    (13,7, "ISA-20098-001",1,D(2009,9,14),1,  8000.00, 172000.00,20000.00,4),
    (14,7, "SIP-20098-001",3,D(2009,9,14),1, 22000.00, 478000.00,    0.00,4),
    (15,8, "ISA-20051-001",1,D(2005,1,22),1, 15000.00, 335000.00,20000.00,4),
    (16,8, "SIP-20051-001",3,D(2005,1,22),1, 28000.00, 652000.00,    0.00,4),
    (17,8, "GIA-20051-001",2,D(2010,4,14),1, 12000.00, 208000.00,    0.00,4),
    (18,9, "ISA-19978-001",1,D(1997,8, 5),1, 20000.00, 500000.00,20000.00,4),
    (19,9, "GIA-19978-001",2,D(1997,8, 5),1, 45000.00,1155000.00,    0.00,4),
    (20,9, "SIP-19978-001",3,D(1999,4,10),1, 18000.00, 362000.00,    0.00,4),
    (21,10,"ISA-20073-001",1,D(2007,3,17),1, 10000.00, 190000.00,20000.00,4),
    (22,10,"SIP-20073-001",3,D(2007,3,17),1, 48000.00,1152000.00,    0.00,4),
    (23,10,"GIA-20073-001",2,D(2012,7,22),1, 18000.00, 402000.00,    0.00,4),
]
account_hdr = ["AccountID","ClientID","AccountNumber","AccountTypeID",
               "OpenedDate","AccountStatusID","CashBalance","TotalMarketValue",
               "ISAAllowanceUsedThisYear","TaxYearID"]

# ── Fund ──────────────────────────────────────────────────────────────────────
# FundTypeID: 1=ETF 2=MutualFund 5=Cash
# AssetClassID: 1=Global Equity 2=UK Equity 5=EM Equity 8=Mixed Asset 14=Cash
# CurrencyID: 1=GBP
fund_rows = [
    (1, "Artisan Partners Global Value GBP Acc",        "GB00B3W3V038",2,1,"Global",         1,0.85,5, 23.29,1.80,12.4,28.6,52.1, 2.8,"Concentrated global value equity fund seeking undervalued companies with durable competitive advantages."),
    (2, "Royal London Global Equity Diversified Z Acc", "GB00B6TKHP24",2,1,"Global",         1,0.72,5, 15.36,2.10,11.8,26.4,48.3, 2.4,"Diversified global equity fund blending quality growth and value styles across all market caps."),
    (3, "JPMorgan Emerging Markets Income C Net Acc",   "GB00B6TLM047",2,5,"Emerging Markets",1,0.95,6, 19.61,2.90, 8.2,18.4,31.7, 1.6,"Targets higher-yielding stocks in emerging market countries, balancing income with long-term capital growth."),
    (4, "Veritas Global Focus GBP C Acc",               "GB00B5VS6B29",2,1,"Global",         1,0.90,5, 22.67,1.60,-3.1,11.2,29.4, 0.8,"High-conviction global equity fund. Benchmark-agnostic, typically 25-35 holdings. Capital preservation focus."),
    (5, "Vanguard FTSE All-World ETF",                  "IE00B3RBWM25",1,1,"Global",         1,0.22,5, 91.24,1.60,18.4,42.1,73.6, 4.2,"Broad global equity ETF tracking the FTSE All-World Index across 3,900+ companies in developed and emerging markets."),
    (6, "iShares Core MSCI World ETF",                  "IE00B4L5Y983",1,1,"Developed Mkts", 1,0.20,5, 78.35,1.40,19.1,44.7,79.2, 4.8,"Low-cost ETF tracking the MSCI World Index — 1,600+ large and mid-cap equities across 23 developed market countries."),
    (7, "Fundsmith Equity Fund T Acc",                  "GB00B4Q5X527",2,1,"Global",         1,1.05,6,622.50,0.00,11.2,18.4,51.3, 2.1,"High-conviction global equity fund. Concentrated portfolio of quality companies with sustainable competitive advantages."),
    (8, "Vanguard LifeStrategy 80% Equity",             "GB00B4PQW151",2,8,"Global",         1,0.22,4,312.18,2.20,13.7,24.9,58.2, 3.1,"Ready-made 80% equity / 20% bond fund. Automatically rebalanced, diversified growth at very low cost."),
    (9, "Royal London UK Equity Income Z Acc",          "GB00B3LRCQ57",2,2,"United Kingdom", 1,0.78,5,215.64,4.20, 8.9,19.2,38.7, 1.8,"Actively managed UK equity income fund. Aims to deliver yield above FTSE All-Share with long-term capital growth."),
    (10,"Cash (GBP)",                                   "CASH-GBP",    5,14,"United Kingdom",1,0.00,1,  1.00,4.50, 4.5, 4.5, 4.5, 4.5,"Sterling cash held on deposit. Interest accrues at prevailing bank base rate."),
]
fund_hdr = ["FundID","FundName","ISIN","FundTypeID","AssetClassID","Geography",
            "CurrencyID","OCF_Pct","KIIDRiskRating","CurrentPrice","EstimatedYield_Pct",
            "Return1yr_Pct","Return3yr_Pct","Return5yr_Pct","ReturnYTD_Pct","Description"]

# ── Holding ───────────────────────────────────────────────────────────────────
LU = D(2026,3,6)
holding_rows = [
    (1, 1,8,  32.04, 281.22,  9010.00,  10000.00,   990.00,+11.0, 220.00,80.0,LU),
    (2, 1,10,2500.00,  1.00,  2500.00,   2500.00,     0.00,  0.0, 112.50,20.0,LU),
    (3, 2,7,  38.64, 699.10, 27020.00,  24070.00, -2950.00,-10.9,   0.00,75.4,LU),
    (4, 2,5,  73.20,  78.40,  5740.00,   6680.00,   940.00,+16.4, 106.88,20.9,LU),
    (5, 2,10,1200.00,  1.00,  1200.00,   1200.00,     0.00,  0.0,  54.00, 3.7,LU),
    (6, 3,8,  38.44, 283.30, 10890.00,  12010.00,  1120.00,+10.3, 264.22,93.8,LU),
    (7, 3,10, 800.00,  1.00,   800.00,    800.00,     0.00,  0.0,  36.00, 6.2,LU),
    (8, 4,6, 320.50,  74.20, 23780.00,  25111.18,  1331.18, +5.6, 351.56,61.9,LU),
    (9, 4,9,  61.82, 185.20, 11450.00,  13332.11,  1882.11,+16.4, 560.15,32.9,LU),
    (10,4,10,3400.00,  1.00,  3400.00,   3400.00,     0.00,  0.0, 153.00, 8.4,LU),
    (11,5,5, 384.80,  72.40, 27860.00,  35082.11,  7222.11,+25.9, 561.31,49.3,LU),
    (12,5,1,1212.00,  19.80, 24000.00,  28227.48,  4227.48,+17.6, 508.09,39.7,LU),
    (13,5,10,4200.00,  1.00,  4200.00,   4200.00,     0.00,  0.0, 189.00, 5.9,LU),
    (14,6,2, 958.40,  14.10, 13510.00,  14726.30,  1216.30, +9.0, 309.25,82.2,LU),
    (15,6,10,1800.00,  1.00,  1800.00,   1800.00,     0.00,  0.0,  81.00,10.1,LU),
    (16,7,6, 482.20,  72.80, 35120.00,  37786.03,  2666.03, +7.6, 529.00,45.7,LU),
    (17,7,7,  42.00, 566.20, 23780.00,  26145.00,  2365.00, +9.9,   0.00,31.6,LU),
    (18,7,9,  82.40, 190.30, 15680.00,  17769.34,  2089.34,+13.3, 746.31,21.5,LU),
    (19,7,10,5000.00,  1.00,  5000.00,   5000.00,     0.00,  0.0, 225.00, 6.1,LU),
    (20,8,1,2840.00,  18.90, 53680.00,  66143.60, 12463.60,+23.2,1190.58,41.6,LU),
    (21,8,3,2240.00,  18.20, 40768.00,  43927.40,  3159.40, +7.7,1273.89,27.6,LU),
    (22,8,2,2980.00,  14.20, 42316.00,  45772.80,  3456.80, +8.2, 960.23,28.8,LU),
    (23,8,10,8000.00,  1.00,  8000.00,   8000.00,     0.00,  0.0, 360.00, 5.0,LU),
    (24,9,5, 214.60,  78.40, 16824.00,  19558.30,  2734.30,+16.3, 313.00,84.3,LU),
    (25,9,10,2000.00,  1.00,  2000.00,   2000.00,     0.00,  0.0,  90.00, 5.1,LU),
    (26,10,5,580.40,  74.20, 43062.00,  52893.26,  9831.26,+22.8, 846.29,57.6,LU),
    (27,10,6,396.80,  74.90, 29710.00,  31081.88,  1371.88, +4.6, 435.15,33.8,LU),
    (28,10,10,4000.00, 1.00,  4000.00,   4000.00,     0.00,  0.0, 180.00, 4.4,LU),
    (29,11,1,3820.00,  19.40, 74110.00,  88927.80, 14817.80,+20.0,1600.70,26.3,LU),
    (30,11,2,5240.00,  14.50, 75980.00,  80486.40,  4506.40, +5.9,1690.21,23.8,LU),
    (31,11,3,3980.00,  18.40, 73240.00,  78146.80,  4906.80, +6.7,2266.26,23.1,LU),
    (32,11,4,2640.00,  22.80, 60192.00,  59848.80,  -343.20, -0.6, 957.58,17.7,LU),
    (33,11,10,12000.00,1.00,  12000.00,  12000.00,     0.00,  0.0, 540.00, 3.5,LU),
    (34,12,1,12430.21, 17.80,221180.00, 289420.00, 68240.00,+30.9,5209.56,34.4,LU),
    (35,12,2,14210.88, 12.42,176540.00, 218340.00, 41800.00,+23.7,4585.14,25.9,LU),
    (36,12,3, 8920.44, 17.06,152200.00, 174890.00, 22690.00,+14.9,5071.81,20.8,LU),
    (37,12,4, 5640.77, 24.20,136490.00, 127890.00, -8600.00, -6.3,2046.24,15.2,LU),
    (38,12,10,31800.00, 1.00, 31800.00,  31800.00,     0.00,  0.0,1431.00, 3.8,LU),
    (39,13,5,  980.20,  80.20, 78575.00,  89334.23, 10759.23,+13.7,1429.35,49.1,LU),
    (40,13,6,  620.40,  75.10, 46601.00,  48621.34,  2020.34, +4.3, 680.70,26.7,LU),
    (41,13,9,  168.40, 192.40, 32400.00,  36304.58,  3904.58,+12.1,1524.79,19.9,LU),
    (42,13,10,8000.00,   1.00,  8000.00,   8000.00,     0.00,  0.0, 360.00, 4.4,LU),
    (43,14,1, 6420.00,  19.20,123264.00, 149476.80, 26212.80,+21.3,2690.58,29.8,LU),
    (44,14,2, 8680.00,  14.40,124992.00, 133284.80,  8292.80, +6.6,2798.98,26.6,LU),
    (45,14,3, 8920.00,  17.60,157000.00, 175058.20, 18058.20,+11.5,5076.69,34.9,LU),
    (46,14,10,22000.00,  1.00, 22000.00,  22000.00,     0.00,  0.0, 990.00, 4.4,LU),
    (47,15,5, 1840.00,  78.40,144256.00, 167801.60, 23545.60,+16.3,2684.83,46.5,LU),
    (48,15,6, 1020.00,  75.60, 77112.00,  79917.00,  2805.00, +3.6,1118.84,22.1,LU),
    (49,15,9,  384.00, 188.20, 72268.00,  82805.76, 10537.76,+14.6,3477.84,22.9,LU),
    (50,15,10,15000.00,  1.00, 15000.00,  15000.00,     0.00,  0.0, 675.00, 4.2,LU),
    (51,16,1, 9840.00,  19.20,188928.00, 229137.60, 40209.60,+21.3,4124.48,33.4,LU),
    (52,16,2,10620.00,  14.40,152928.00, 163027.20, 10099.20, +6.6,3423.57,23.8,LU),
    (53,16,3, 9840.00,  17.20,169248.00, 193030.40, 23782.40,+14.1,5597.88,28.1,LU),
    (54,16,4, 3240.00,  22.60, 73224.00,  73451.80,   227.80, +0.3,1175.23,10.7,LU),
    (55,16,10,28000.00,  1.00, 28000.00,  28000.00,     0.00,  0.0,1260.00, 4.1,LU),
    (56,17,7,  184.20, 568.40,104657.00, 114664.50, 10007.50, +9.6,   0.00,52.1,LU),
    (57,17,6,  620.00,  74.80, 46376.00,  48577.00,  2201.00, +4.7, 680.08,22.1,LU),
    (58,17,10,12000.00,  1.00, 12000.00,  12000.00,     0.00,  0.0, 540.00, 5.5,LU),
    (59,18,5, 2480.00,  79.60,197408.00, 226275.20, 28867.20,+14.6,3620.40,43.5,LU),
    (60,18,6, 1620.00,  75.40,122148.00, 126927.00,  4779.00, +3.9,1777.00,24.4,LU),
    (61,18,9,  684.00, 188.80,129139.00, 147497.76, 18358.76,+14.2,6194.96,28.3,LU),
    (62,18,10,20000.00,  1.00, 20000.00,  20000.00,     0.00,  0.0, 900.00, 3.8,LU),
    (63,19,1,18420.00,  18.80,346296.00, 428923.80, 82627.80,+23.9,7720.63,35.8,LU),
    (64,19,2,18200.00,  14.20,258440.00, 279279.20, 20839.20, +8.1,5864.86,23.3,LU),
    (65,19,3,14800.00,  17.40,257520.00, 290228.00, 32708.00,+12.7,8416.61,24.2,LU),
    (66,19,7,  384.00, 580.00,222720.00, 239040.00, 16320.00, +7.3,   0.00,19.9,LU),
    (67,19,4, 2480.00,  23.20, 57536.00,  56221.60, -1314.40, -2.3, 899.55, 4.7,LU),
    (68,19,10,45000.00,  1.00, 45000.00,  45000.00,     0.00,  0.0,2025.00, 3.8,LU),
    (69,20,5, 1980.00,  80.40,159192.00, 180655.20, 21463.20,+13.5,2890.48,47.9,LU),
    (70,20,6, 1440.00,  75.80,109152.00, 112824.00,  3672.00, +3.4,1579.54,29.9,LU),
    (71,20,10,18000.00,  1.00, 18000.00,  18000.00,     0.00,  0.0, 810.00, 4.8,LU),
    (72,21,5, 1240.00,  79.60, 98704.00, 113137.60, 14433.60,+14.6,1810.20,56.8,LU),
    (73,21,7,  116.00, 560.00, 64960.00,  72210.00,  7250.00,+11.2,   0.00,36.3,LU),
    (74,21,10,10000.00,  1.00, 10000.00,  10000.00,     0.00,  0.0, 450.00, 5.0,LU),
    (75,22,1,15840.00,  19.20,304128.00, 368841.60, 64713.60,+21.3,6638.95,30.8,LU),
    (76,22,2,14280.00,  14.40,205632.00, 219250.80, 13618.80, +6.6,4604.27,18.3,LU),
    (77,22,3,13800.00,  17.40,240120.00, 270618.00, 30498.00,+12.7,7847.92,22.6,LU),
    (78,22,4, 5640.00,  23.20,130848.00, 127801.80, -3046.20, -2.3,2044.83,10.7,LU),
    (79,22,5, 1620.00,  79.60,128952.00, 147808.80, 18856.80,+14.6,2364.94,12.3,LU),
    (80,22,10,48000.00,  1.00, 48000.00,  48000.00,     0.00,  0.0,2160.00, 4.0,LU),
    (81,23,6, 2820.00,  74.80,210936.00, 220947.00, 10011.00, +4.7,3093.26,52.5,LU),
    (82,23,9,  864.00, 190.40,164506.00, 186313.00, 21807.00,+13.3,7825.15,44.2,LU),
    (83,23,10,18000.00,  1.00, 18000.00,  18000.00,     0.00,  0.0, 810.00, 4.3,LU),
]
holding_hdr = ["HoldingID","AccountID","FundID","Units","AverageCostPrice",
               "TotalCostBasis","CurrentMarketValue","UnrealisedGainLoss",
               "UnrealisedGainLossPct","EstimatedAnnualYield","PctOfPortfolio","LastUpdated"]

# ── Transaction ───────────────────────────────────────────────────────────────
# TransactionTypeID: 1=Buy 2=Sell 3=Dividend 4=Deposit  TransactionStatusID: 1=Settled
txn_rows = [
    (1, 1,8, 4,D(2024,1,15),D(2024,1,16),   0.00,   0.00, 5000.00,0.00, 5000.00,"DEP-2401-001",1),
    (2, 1,8, 1,D(2024,1,18),D(2024,1,20),  16.00, 280.50, 4488.00,0.00, 4488.00,"TRD-2401-001",1),
    (3, 1,8, 4,D(2024,6,15),D(2024,6,16),   0.00,   0.00, 5000.00,0.00, 5000.00,"DEP-2406-001",1),
    (4, 1,8, 1,D(2024,6,18),D(2024,6,20),  16.04, 281.22, 4510.00,0.00, 4510.00,"TRD-2406-001",1),
    (5, 1,8, 3,D(2025,3, 1),D(2025,3, 3),   0.00,   0.00,  186.00,0.00,  186.00,"DIV-2503-001",1),
    (6, 2,7, 1,D(2023,4,12),D(2023,4,14),  24.00, 648.20,15556.80,0.00,15556.80,"TRD-2304-001",1),
    (7, 2,5, 1,D(2023,4,12),D(2023,4,14),  48.00,  79.10, 3796.80,0.00, 3796.80,"TRD-2304-002",1),
    (8, 2,7, 1,D(2024,9,20),D(2024,9,22),  14.64, 639.40, 9360.50,0.00, 9360.50,"TRD-2409-001",1),
    (9, 2,5, 1,D(2024,9,20),D(2024,9,22),  25.20,  79.20, 1995.84,0.00, 1995.84,"TRD-2409-002",1),
    (10,2,5, 3,D(2025,3, 1),D(2025,3, 3),   0.00,   0.00,   96.48,0.00,   96.48,"DIV-2503-002",1),
    (11,3,8, 1,D(2023,4,12),D(2023,4,14),  38.44, 283.30,10890.00,0.00,10890.00,"TRD-2304-003",1),
    (12,3,8, 3,D(2025,3, 1),D(2025,3, 3),   0.00,   0.00,  216.40,0.00,  216.40,"DIV-2503-003",1),
    (13,4,6, 1,D(2022,3,15),D(2022,3,17), 200.00,  74.20,14840.00,0.00,14840.00,"TRD-2203-001",1),
    (14,4,6, 1,D(2023,9,10),D(2023,9,12), 120.50,  74.20, 8941.10,0.00, 8941.10,"TRD-2309-001",1),
    (15,4,9, 1,D(2023,9,10),D(2023,9,12),  61.82, 185.20,11450.00,0.00,11450.00,"TRD-2309-002",1),
    (16,4,9, 3,D(2025,3, 1),D(2025,3, 3),   0.00,   0.00,  463.00,0.00,  463.00,"DIV-2503-004",1),
    (17,5,5, 1,D(2021,6,14),D(2021,6,16), 240.00,  68.40,16416.00,0.00,16416.00,"TRD-2106-001",1),
    (18,5,5, 1,D(2023,3, 8),D(2023,3,10), 144.80,  77.40,11207.52,0.00,11207.52,"TRD-2303-001",1),
    (19,5,1, 1,D(2022,11,14),D(2022,11,16),1212.00,19.80,24000.00,0.00,24000.00,"TRD-2211-001",1),
    (20,5,5, 3,D(2025,3, 1),D(2025,3, 3),   0.00,   0.00,  450.40,0.00,  450.40,"DIV-2503-005",1),
    (21,5,1, 3,D(2025,3, 1),D(2025,3, 3),   0.00,   0.00,  432.20,0.00,  432.20,"DIV-2503-006",1),
    (22,12,2,1,D(2026,3, 2),D(2026,3, 4), 842.40,  14.82,12480.00,0.00,12480.00,"TRD-2603-001",1),
    (23,12,1,3,D(2026,2,28),D(2026,3, 1),  0.00,   0.00, 1840.00,0.00, 1840.00,"DIV-2602-001",1),
    (24,12,3,2,D(2026,2,14),D(2026,2,16), 420.10,  19.61, 8241.16,0.00, 8241.16,"TRD-2602-001",1),
    (25,12,2,3,D(2026,1,31),D(2026,2, 2),  0.00,   0.00, 2210.00,0.00, 2210.00,"DIV-2601-001",1),
    (26,12,1,1,D(2026,1,15),D(2026,1,17), 610.00,  23.10,14091.00,0.00,14091.00,"TRD-2601-001",1),
    (27,12,1,1,D(2024,12,10),D(2024,12,12),840.00, 21.40,17976.00,0.00,17976.00,"TRD-2412-001",1),
    (28,13,5,1,D(2024,2,14),D(2024,2,16), 420.00,  79.10,33222.00,0.00,33222.00,"TRD-2402-001",1),
    (29,13,9,1,D(2024,2,14),D(2024,2,16), 168.40, 192.40,32400.00,0.00,32400.00,"TRD-2402-002",1),
    (30,13,5,1,D(2025,4,10),D(2025,4,12), 560.20,  81.30,45544.26,0.00,45544.26,"TRD-2504-001",1),
    (31,13,9,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 1420.80,0.00, 1420.80,"DIV-2602-002",1),
    (32,14,1,1,D(2023,6,20),D(2023,6,22),3840.00,  18.80,72192.00,0.00,72192.00,"TRD-2306-001",1),
    (33,14,3,1,D(2023,6,20),D(2023,6,22),5640.00,  17.20,96988.00,0.00,96988.00,"TRD-2306-002",1),
    (34,14,2,1,D(2024,9,16),D(2024,9,18),5380.00,  14.20,76396.00,0.00,76396.00,"TRD-2409-003",1),
    (35,14,3,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 4631.04,0.00, 4631.04,"DIV-2602-003",1),
    (36,15,5,1,D(2022,4,20),D(2022,4,22),1020.00,  76.40,77928.00,0.00,77928.00,"TRD-2204-001",1),
    (37,15,5,1,D(2024,10,14),D(2024,10,16),820.00, 80.80,66256.00,0.00,66256.00,"TRD-2410-001",1),
    (38,15,9,1,D(2024,10,14),D(2024,10,16),384.00,188.20,72268.00,0.00,72268.00,"TRD-2410-002",1),
    (39,15,9,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 2880.00,0.00, 2880.00,"DIV-2602-004",1),
    (40,16,1,1,D(2022,6, 8),D(2022,6,10),6840.00,  18.80,128592.00,0.00,128592.00,"TRD-2206-001",1),
    (41,16,3,1,D(2022,6, 8),D(2022,6,10),7200.00,  17.20,123840.00,0.00,123840.00,"TRD-2206-002",1),
    (42,16,2,1,D(2024,3,18),D(2024,3,20),7380.00,  14.20,104796.00,0.00,104796.00,"TRD-2403-001",1),
    (43,16,1,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 3516.00,0.00, 3516.00,"DIV-2602-005",1),
    (44,18,5,1,D(2023,5,12),D(2023,5,14),1480.00,  77.80,115144.00,0.00,115144.00,"TRD-2305-001",1),
    (45,18,5,1,D(2025,4, 8),D(2025,4,10),1000.00,  82.26,82264.00,0.00,82264.00,"TRD-2504-002",1),
    (46,18,9,1,D(2024,7,22),D(2024,7,24), 684.00, 188.80,129139.00,0.00,129139.00,"TRD-2407-001",1),
    (47,18,9,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 5390.40,0.00, 5390.40,"DIV-2602-006",1),
    (48,22,1,1,D(2021,8,16),D(2021,8,18),9840.00,  17.40,171216.00,0.00,171216.00,"TRD-2108-001",1),
    (49,22,3,1,D(2021,8,16),D(2021,8,18),8800.00,  16.80,147840.00,0.00,147840.00,"TRD-2108-002",1),
    (50,22,2,1,D(2023,11,20),D(2023,11,22),9280.00,14.20,131776.00,0.00,131776.00,"TRD-2311-001",1),
    (51,22,1,1,D(2025,2,10),D(2025,2,12),6000.00,  19.80,118800.00,0.00,118800.00,"TRD-2502-001",1),
    (52,22,1,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 5702.40,0.00, 5702.40,"DIV-2602-007",1),
    (53,22,3,3,D(2026,2, 1),D(2026,2, 3),  0.00,   0.00, 7159.62,0.00, 7159.62,"DIV-2602-008",1),
]
txn_hdr = ["TransactionID","AccountID","FundID","TransactionTypeID",
           "TransactionDate","SettlementDate","Units","UnitPrice",
           "GrossAmount","Charges","NetAmount","Reference","TransactionStatusID"]

# ── Document ──────────────────────────────────────────────────────────────────
# DocumentCategoryID: 1=Statement 2=Tax 3=Correspondence 4=Report 5=Form
# AccountID=0 means applies to all accounts
doc_rows = [
    (1, 1, 1,5,"ISA Account Opening Pack",                     D(2024,1,15), 420,False),
    (2, 1, 1,1,"Portfolio Statement — February 2026",          D(2026,2, 6), 148,True),
    (3, 1, 1,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (4, 2, 2,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 210,True),
    (5, 2, 0,2,"CGT Summary Report 2025-26",                   D(2026,3, 6), 188,True),
    (6, 2, 2,3,"Annual Review Letter 2026",                    D(2026,3, 4), 245,False),
    (7, 2, 2,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (8, 2, 3,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (9, 3, 4,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 312,True),
    (10,3, 0,3,"Annual Review Letter 2026",                    D(2026,2,20), 310,False),
    (11,3, 4,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (12,3, 5,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (13,3, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (14,4, 7,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 398,True),
    (15,4, 0,3,"Annual Review Letter 2026",                    D(2026,3, 4), 312,True),
    (16,4, 0,2,"CGT Summary Report 2025-26",                   D(2026,3, 6), 248,False),
    (17,4, 7,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (18,4, 8,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (19,4, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (20,5,10,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 420,True),
    (21,5, 0,3,"Annual Review Letter 2026",                    D(2026,3, 4), 312,False),
    (22,5,10,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (23,5,11,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (24,5, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (25,5, 0,2,"Dividend Voucher — Artisan Partners Q4",       D(2026,2,28),  84,False),
    (26,6,12,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 420,True),
    (27,6, 0,2,"CGT Summary Report 2025-26",                   D(2026,3, 6), 248,True),
    (28,6, 0,3,"Annual Review Letter 2026",                    D(2026,3, 4), 312,True),
    (29,6,12,2,"Dividend Voucher — Artisan Partners Q4",       D(2026,2,28),  84,False),
    (30,6,12,1,"Portfolio Statement — February 2026",          D(2026,2, 6), 398,False),
    (31,6, 0,3,"Regulatory Notice — MIFID II Update",          D(2026,2, 1),  96,False),
    (32,6,12,1,"Portfolio Statement — January 2026",           D(2026,1, 6), 385,False),
    (33,6, 0,4,"Valuation Report Q4 2025",                     D(2025,12,31),312,False),
    (34,6, 0,4,"Quarterly Investment Report — Q3 2025",        D(2025,10,15),784,False),
    (35,6,12,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (36,6, 0,2,"Annual CGT Statement 2024-25",                 D(2025,4, 6), 195,False),
    (37,6, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (38,6, 0,5,"Investment Management Agreement",              D(2024,1,12), 328,False),
    (39,7,13,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 520,True),
    (40,7, 0,3,"Annual Review Letter 2026",                    D(2026,3, 4), 312,True),
    (41,7, 0,2,"CGT Summary Report 2025-26",                   D(2026,3, 6), 248,False),
    (42,7,13,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (43,7,14,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (44,7, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (45,8,15,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 684,True),
    (46,8, 0,3,"Annual Review Letter 2026",                    D(2026,3, 4), 312,True),
    (47,8, 0,2,"CGT Summary Report 2025-26",                   D(2026,3, 6), 248,False),
    (48,8,15,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (49,8,16,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (50,8, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (51,8, 0,4,"Valuation Report Q4 2025",                     D(2025,12,31),312,False),
    (52,9,18,1,"Portfolio Statement — March 2026",             D(2026,3, 6), 840,True),
    (53,9, 0,3,"Annual Review Letter 2026",                    D(2026,3, 4), 312,True),
    (54,9, 0,2,"CGT Summary Report 2025-26",                   D(2026,3, 6), 248,False),
    (55,9,18,5,"ISA Subscription Confirmation 2025-26",        D(2025,4,10), 142,False),
    (56,9,20,2,"SIPP Annual Allowance Statement 2024-25",      D(2025,4, 6), 189,False),
    (57,9, 0,4,"Suitability Report — Annual Review 2025",      D(2025,3,15),1248,False),
    (58,9, 0,4,"Valuation Report Q4 2025",                     D(2025,12,31),312,False),
    (59,9, 0,4,"Quarterly Investment Report — Q3 2025",        D(2025,10,15),784,False),
    (60,10,21,1,"Portfolio Statement — March 2026",            D(2026,3, 6), 720,True),
    (61,10, 0,3,"Annual Review Letter 2026",                   D(2026,3, 4), 312,True),
    (62,10, 0,2,"CGT Summary Report 2025-26",                  D(2026,3, 6), 248,False),
    (63,10,21,5,"ISA Subscription Confirmation 2025-26",       D(2025,4,10), 142,False),
    (64,10,22,2,"SIPP Annual Allowance Statement 2024-25",     D(2025,4, 6), 189,False),
    (65,10, 0,4,"Suitability Report — Annual Review 2025",     D(2025,3,15),1248,False),
    (66,10, 0,4,"Valuation Report Q4 2025",                    D(2025,12,31),312,False),
    (67,10, 0,2,"Dividend Voucher — Artisan Partners Q4",      D(2026,2,28),  84,False),
]
doc_hdr = ["DocumentID","ClientID","AccountID","DocumentCategoryID",
           "DocumentName","DocumentDate","FileSizeKB","IsUnread"]

# ── ValuationSnapshot ─────────────────────────────────────────────────────────
val_hdr = ["ValuationID","ClientID","SnapshotDate","TotalPortfolioValue",
           "TotalInvestedValue","TotalCashValue","TotalCostBasis",
           "TotalGainLoss","TotalGainLossPct"]
quarters = [D(2024,3,31),D(2024,6,30),D(2024,9,30),D(2024,12,31),
            D(2025,3,31),D(2025,6,30),D(2025,9,30),D(2025,12,31)]
final_v = {1:12500,2:45200,3:128000,4:284000,5:425000,
           6:842340,7:680000,8:1250000,9:2100000,10:1820000}
cost_b  = {1:9010,2:38950,3:107000,4:234000,5:360000,
           6:718210,7:548000,8:1038000,9:1600000,10:1450000}
sf      = {1:0.80,2:0.72,3:0.68,4:0.71,5:0.72,
           6:0.74,7:0.70,8:0.72,9:0.74,10:0.73}
val_rows, vid = [], 1
for cid in range(1,11):
    n = len(quarters)
    for i,qd in enumerate(quarters):
        f   = sf[cid] + (1.0-sf[cid])*(i/(n-1))*0.92
        tv  = round(final_v[cid]*f,-2)
        cb  = round(cost_b[cid]*(0.85+0.15*(i/(n-1))),-2)
        cash= round(tv*0.04,-2)
        gl  = tv-cb
        glp = round((gl/cb)*100,1) if cb else 0.0
        val_rows.append((vid,cid,qd,tv,tv-cash,cash,cb,gl,glp))
        vid+=1


# =============================================================================
# BUILD WORKBOOK
# =============================================================================
wb = openpyxl.Workbook()
del wb[wb.sheetnames[0]]

# ── README ────────────────────────────────────────────────────────────────────
ws0 = wb.create_sheet("README")
ws0["A1"] = "Aurelius Wealth Management — Seed Data (v2 Fully Relational)"
ws0["A1"].font = Font(bold=True,size=13,name="Calibri",color="C49A38")
ws0["A2"] = "Generated: " + str(datetime.date.today())
ws0["A2"].font = Font(italic=True,size=10,name="Calibri",color="888888")
ws0["A4"] = "LOOKUP TABLES (teal headers) — import first"
ws0["A4"].font = Font(bold=True,size=11,name="Calibri")
lookup_index = [
    ("LKP_Title","Salutation titles (Mr, Mrs, Ms, Dr …)"),
    ("LKP_Gender","Gender options"),
    ("LKP_MaritalStatus","Marital status options"),
    ("LKP_Nationality","Nationalities and home countries"),
    ("LKP_Country","Countries with ISO-2 and ISO-3 codes"),
    ("LKP_Branch","Aurelius office branches"),
    ("LKP_AdviserRole","Adviser role levels"),
    ("LKP_ClientTier","Client tier bands with minimum portfolio values"),
    ("LKP_KYCStatus","KYC verification status options"),
    ("LKP_KYCDocType","Acceptable KYC identity document types"),
    ("LKP_AccountType","Investment wrapper types with annual allowances"),
    ("LKP_AccountStatus","Account status options"),
    ("LKP_FundType","Fund / instrument type classifications"),
    ("LKP_AssetClass","Asset class taxonomy"),
    ("LKP_Currency","Currencies with ISO codes and symbols"),
    ("LKP_TransactionType","Transaction types with credit/debit flag"),
    ("LKP_TransactionStatus","Transaction settlement status options"),
    ("LKP_DocumentCategory","Document category taxonomy"),
    ("LKP_AddressType","Address type options"),
    ("LKP_ContactMethod","Preferred contact method options"),
    ("LKP_NotifChannel","Notification channel options"),
    ("LKP_NotifCategory","Notification category options"),
    ("LKP_LifeStage","Life stage options for investment profile"),
    ("LKP_RiskTolerance","Risk tolerance levels"),
    ("LKP_WealthPriority","Wealth objective priority options"),
    ("LKP_CapitalRange","Investable capital range bands"),
    ("LKP_Element","Five-element investment personality types"),
    ("LKP_AppTheme","UI appearance theme options"),
    ("LKP_TaxYear","Tax years with allowances"),
]
for i,(s,d) in enumerate(lookup_index,5):
    ws0[f"A{i}"] = s; ws0[f"B{i}"] = d
    ws0[f"A{i}"].font = Font(name="Calibri",size=10,bold=True,color="7FB8D4")
    ws0[f"B{i}"].font = Font(name="Calibri",size=10)

r = 5+len(lookup_index)+1
ws0[f"A{r}"] = "ENTITY TABLES (navy headers) — import after lookups"
ws0[f"A{r}"].font = Font(bold=True,size=11,name="Calibri")
entity_index = [
    ("Adviser","5 wealth managers across 4 UK branches"),
    ("Client","10 UK clients — all categorical fields use FK IDs"),
    ("ClientContact","Email, phone numbers, preferred contact method (FK)"),
    ("ClientAddress","Registered and correspondence addresses — AddressTypeID + CountryID FKs"),
    ("ClientPreferences","Consents, paperless, AppThemeID FK"),
    ("NotifPreferences","30 rows — ChannelID FK, boolean columns per category"),
    ("InvestmentProfile","Know Your Element results — all categorical fields as FK IDs"),
    ("Account","23 accounts — AccountTypeID + AccountStatusID + TaxYearID FKs"),
    ("Fund","10 funds — FundTypeID + AssetClassID + CurrencyID FKs"),
    ("Holding","83 positions across all accounts"),
    ("Transaction","53 transactions — TransactionTypeID + TransactionStatusID FKs"),
    ("Document","67 documents — DocumentCategoryID FK"),
    ("ValuationSnapshot","80 quarterly valuations (Q1 2024 – Q4 2025)"),
]
for i,(s,d) in enumerate(entity_index, r+1):
    ws0[f"A{i}"] = s; ws0[f"B{i}"] = d
    ws0[f"A{i}"].font = Font(name="Calibri",size=10,bold=True)
    ws0[f"B{i}"].font = Font(name="Calibri",size=10)
ws0.column_dimensions["A"].width = 26
ws0.column_dimensions["B"].width = 62

# ── Lookup sheets ─────────────────────────────────────────────────────────────
sheet(wb,"LKP_Title",          ["TitleID","Title"],                                                lkp_title,      True)
sheet(wb,"LKP_Gender",         ["GenderID","Gender"],                                              lkp_gender,     True)
sheet(wb,"LKP_MaritalStatus",  ["MaritalStatusID","MaritalStatus"],                                lkp_marital,    True)
sheet(wb,"LKP_Nationality",    ["NationalityID","Nationality","Country"],                          lkp_nationality,True)
sheet(wb,"LKP_Country",        ["CountryID","CountryName","ISO2","ISO3"],                          lkp_country,    True)
sheet(wb,"LKP_Branch",         ["BranchID","BranchName","City","AddressLine1","Postcode"],         lkp_branch,     True)
sheet(wb,"LKP_AdviserRole",    ["RoleID","RoleName","RoleLevel"],                                  lkp_adviser_role,True)
sheet(wb,"LKP_ClientTier",     ["TierID","TierName","MinPortfolioValue","Description"],            lkp_tier,       True)
sheet(wb,"LKP_KYCStatus",      ["KYCStatusID","StatusName","Description"],                        lkp_kyc_status, True)
sheet(wb,"LKP_KYCDocType",     ["DocTypeID","DocumentType"],                                       lkp_kyc_doc,    True)
sheet(wb,"LKP_AccountType",    ["AccountTypeID","AccountTypeCode","AccountTypeName","AnnualAllowance","Description"],lkp_account_type,True)
sheet(wb,"LKP_AccountStatus",  ["AccountStatusID","StatusName"],                                   lkp_account_status,True)
sheet(wb,"LKP_FundType",       ["FundTypeID","FundType","Description"],                            lkp_fund_type,  True)
sheet(wb,"LKP_AssetClass",     ["AssetClassID","AssetClass"],                                      lkp_asset_class,True)
sheet(wb,"LKP_Currency",       ["CurrencyID","CurrencyCode","CurrencyName","Symbol"],              lkp_currency,   True)
sheet(wb,"LKP_TransactionType",["TxnTypeID","TxnType","IsCredit","Description"],                   lkp_txn_type,   True)
sheet(wb,"LKP_TransactionStatus",["TxnStatusID","StatusName"],                                     lkp_txn_status, True)
sheet(wb,"LKP_DocumentCategory",["DocCatID","Category","Description"],                             lkp_doc_cat,    True)
sheet(wb,"LKP_AddressType",    ["AddrTypeID","AddressType","Description"],                         lkp_addr_type,  True)
sheet(wb,"LKP_ContactMethod",  ["ContactMethodID","Method"],                                       lkp_contact_method,True)
sheet(wb,"LKP_NotifChannel",   ["ChannelID","Channel"],                                            lkp_notif_channel,True)
sheet(wb,"LKP_NotifCategory",  ["CategoryID","Category","Description"],                            lkp_notif_category,True)
sheet(wb,"LKP_LifeStage",      ["LifeStageID","StageName","StageDescription"],                     lkp_life_stage, True)
sheet(wb,"LKP_RiskTolerance",  ["RiskToleranceID","RiskName","RiskDescription"],                   lkp_risk,       True)
sheet(wb,"LKP_WealthPriority", ["WealthPriorityID","PriorityName"],                                lkp_priority,   True)
sheet(wb,"LKP_CapitalRange",   ["CapitalRangeID","RangeName","RangeMin","RangeMax","Description"], lkp_capital_range,True)
sheet(wb,"LKP_Element",        ["ElementID","ElementName","HexColour","Description"],              lkp_element,    True)
sheet(wb,"LKP_AppTheme",       ["ThemeID","ThemeName"],                                            lkp_theme,      True)
sheet(wb,"LKP_TaxYear",        ["TaxYearID","TaxYearLabel","StartDate","EndDate","ISAAllowance","SIPPAllowance","CGTAllowance"],lkp_tax_year,True)

# ── Entity sheets ─────────────────────────────────────────────────────────────
sheet(wb,"Adviser",          adviser_hdr,  adviser_rows)
sheet(wb,"Client",           client_hdr,   client_rows)
sheet(wb,"ClientContact",    contact_hdr,  contact_rows)
sheet(wb,"ClientAddress",    address_hdr,  address_rows)
sheet(wb,"ClientPreferences",pref_hdr,     pref_rows)
sheet(wb,"NotifPreferences", notif_hdr,    notif_rows)
sheet(wb,"InvestmentProfile",profile_hdr,  profile_rows)
sheet(wb,"Account",          account_hdr,  account_rows)
sheet(wb,"Fund",             fund_hdr,     fund_rows)
sheet(wb,"Holding",          holding_hdr,  holding_rows)
sheet(wb,"Transaction",      txn_hdr,      txn_rows)
sheet(wb,"Document",         doc_hdr,      doc_rows)
sheet(wb,"ValuationSnapshot",val_hdr,      val_rows)

wb.save(OUT_FILE)
total_rows = sum(len(r) for r in [adviser_rows,client_rows,contact_rows,address_rows,
    pref_rows,notif_rows,profile_rows,account_rows,fund_rows,holding_rows,
    txn_rows,doc_rows,val_rows])
lkp_rows = sum(len(r) for r in [lkp_title,lkp_gender,lkp_marital,lkp_nationality,
    lkp_country,lkp_branch,lkp_adviser_role,lkp_tier,lkp_kyc_status,lkp_kyc_doc,
    lkp_account_type,lkp_account_status,lkp_fund_type,lkp_asset_class,lkp_currency,
    lkp_txn_type,lkp_txn_status,lkp_doc_cat,lkp_addr_type,lkp_contact_method,
    lkp_notif_channel,lkp_notif_category,lkp_life_stage,lkp_risk,lkp_priority,
    lkp_capital_range,lkp_element,lkp_theme,lkp_tax_year])
print(f"Saved: {OUT_FILE}")
print(f"  Sheets total : {len(wb.sheetnames)}  (1 README + 29 lookup + 13 entity)")
print(f"  Lookup rows  : {lkp_rows}")
print(f"  Entity rows  : {total_rows}")
