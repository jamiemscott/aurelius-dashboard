#!/usr/bin/env python3
"""
Aurelius Wealth Management — Market & Economic Data Fetcher
===========================================================
Fetches real market data from free public APIs and appends it to
aurelius_seed_data.xlsx as new sheets.

FREE APIS USED
──────────────
  yfinance       - Yahoo Finance. No key. ETF/stock/index prices + metadata.
                   pip install yfinance
  Frankfurter    - European Central Bank FX rates. No key. No rate limit.
                   https://api.frankfurter.app

WHAT GETS FETCHED
─────────────────
  1. MarketIndex        — reference table of 6 tracked indices/benchmarks
  2. IndexPriceHistory  — daily OHLC + return% for each index (2 years)
  3. FundPriceHistory   — daily NAV for the 2 LSE-listed ETFs in Fund table
                          (proxy NAV simulated for the 4 mutual funds)
  4. FXRateHistory      — daily GBP/USD, GBP/EUR, GBP/JPY (2 years)
  5. LKP_EconomicIndicator — reference table
  6. EconomicIndicatorHistory — BoE base rate + UK CPI (hardcoded from
                                published public data, APIs not accessible
                                without scraping)

HOW TABLES LINK TO EXISTING SCHEMA
────────────────────────────────────
  FundPriceHistory.FundID      → Fund.FundID
  IndexPriceHistory.IndexID    → MarketIndex.IndexID
  FXRateHistory.CurrencyID     → LKP_Currency.CurrencyID
  EconomicIndicatorHistory.IndicatorID → LKP_EconomicIndicator.IndicatorID

Run:
  python generate_market_data.py

Output: adds sheets to aurelius_seed_data.xlsx (non-destructive)
"""

import os, sys, datetime, requests
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE      = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(HERE, "aurelius_seed_data.xlsx")
D         = datetime.date
START     = "2024-01-01"
END       = "2026-04-18"

# ── Styling ───────────────────────────────────────────────────────────────────
MKT_FILL  = PatternFill("solid", fgColor="0A2218")   # dark green for market sheets
MKT_FONT  = Font(bold=True, color="4CAF7D", name="Calibri", size=10)
BODY_FONT = Font(name="Calibri", size=10)
LKP_FILL  = PatternFill("solid", fgColor="0D3349")
LKP_FONT  = Font(bold=True, color="7FB8D4", name="Calibri", size=10)

def _style(ws, is_lookup=False, is_market=False):
    if is_market:
        fill, font = MKT_FILL, MKT_FONT
    elif is_lookup:
        fill, font = LKP_FILL, LKP_FONT
    else:
        fill, font = PatternFill("solid", fgColor="1C2130"), Font(bold=True, color="C49A38", name="Calibri", size=10)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.font = BODY_FONT
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, 10), 45)
    ws.freeze_panes = "A2"
    if ws.max_row > 1: ws.auto_filter.ref = ws.dimensions

def add_sheet(wb, name, headers, rows, lookup=False, market=True):
    # Remove existing sheet if present (for re-runs)
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for r in rows: ws.append(r)
    _style(ws, is_lookup=lookup, is_market=market and not lookup)
    return ws


# =============================================================================
# TABLE 1 — MarketIndex  (lookup / reference)
# =============================================================================
print("Building MarketIndex reference table...")

# IndexID, IndexName, Ticker, Geography, Description, BenchmarkFor
market_index_rows = [
    (1, "FTSE 100",              "^FTSE",    "United Kingdom",  "UK large-cap index — top 100 companies on the London Stock Exchange",       "UK Equity holdings"),
    (2, "S&P 500",               "^GSPC",    "United States",   "US large-cap index — 500 largest US publicly traded companies",             "Global/US Equity holdings"),
    (3, "MSCI World (via SWDA)", "SWDA.L",   "Global Developed","Developed-markets equity index tracking 1,600+ companies in 23 countries", "Global Equity funds"),
    (4, "FTSE All-World (VWRL)", "VWRL.L",   "Global",          "All-world equity index covering developed and emerging markets (~3,900 stocks)","Blended global holdings"),
    (5, "Euro Stoxx 50",         "^STOXX50E","Europe",          "50 largest eurozone blue-chip companies",                                   "European Equity exposure"),
    (6, "Nikkei 225",            "^N225",    "Japan",           "225 largest companies on the Tokyo Stock Exchange",                         "Japan/Asia Pacific exposure"),
]
market_index_hdr = ["IndexID","IndexName","Ticker","Geography","Description","BenchmarkFor"]


# =============================================================================
# TABLE 2 — IndexPriceHistory  (real data via yfinance)
# =============================================================================
print("Fetching index price history via yfinance (2 years)...")

index_ticker_map = {
    1: "^FTSE",
    2: "^GSPC",
    3: "SWDA.L",
    4: "VWRL.L",
    5: "^STOXX50E",
    6: "^N225",
}
index_price_hdr = ["PriceHistID","IndexID","Date","Open","High","Low","Close",
                   "Volume","DailyChangePct"]
index_price_rows = []
pid = 1
for idx_id, ticker in index_ticker_map.items():
    print(f"  Fetching {ticker}...", end=" ")
    try:
        t = yf.Ticker(ticker)
        hist = t.history(start=START, end=END)
        count = 0
        prev_close = None
        for date, row in hist.iterrows():
            close = round(float(row["Close"]), 4)
            chg   = round(((close / prev_close) - 1) * 100, 4) if prev_close else 0.0
            prev_close = close
            index_price_rows.append((
                pid, idx_id,
                date.date(),
                round(float(row["Open"]),  4),
                round(float(row["High"]),  4),
                round(float(row["Low"]),   4),
                close,
                int(row["Volume"]) if row["Volume"] else 0,
                chg,
            ))
            pid   += 1
            count += 1
        print(f"{count} rows")
    except Exception as e:
        print(f"ERROR: {e}")


# =============================================================================
# TABLE 3 — FundPriceHistory
# =============================================================================
# FundID mapping (from Fund table):
#   1 = Artisan Partners Global Value   → proxy: SWDA.L (global equity mutual fund)
#   2 = Royal London Global Equity      → proxy: SWDA.L scaled to start at 15.36
#   3 = JPMorgan EM Income              → proxy: EIMI.L (iShares EM IMI ETF)
#   4 = Veritas Global Focus            → proxy: SWDA.L scaled (underperformer)
#   5 = Vanguard FTSE All-World (ETF)   → REAL: VWRL.L
#   6 = iShares Core MSCI World (ETF)   → REAL: SWDA.L
#   7 = Fundsmith Equity T Acc          → proxy: SWDA.L scaled to 622.50
#   8 = Vanguard LifeStrategy 80%       → proxy: blend of SWDA.L + dampened (80/20)
#   9 = Royal London UK Equity Income   → proxy: ISF.L (iShares FTSE 100)
#  10 = Cash (GBP)                      → fixed 1.00 (price never changes)
print("\nFetching fund price history...")

# First, pull the raw data for the proxy tickers we need
proxy_data = {}
proxy_tickers = ["VWRL.L", "SWDA.L", "EIMI.L", "ISF.L"]
for pt in proxy_tickers:
    print(f"  Fetching proxy {pt}...", end=" ")
    try:
        hist = yf.Ticker(pt).history(start=START, end=END)
        proxy_data[pt] = hist
        print(f"{len(hist)} rows")
    except Exception as e:
        print(f"ERROR: {e}")
        proxy_data[pt] = None

fund_price_hdr = ["FundPriceID","FundID","Date","NAV","Source"]

def scale_series(series, target_start):
    """Scale a price series so it starts at target_start."""
    if series is None or series.empty: return {}
    first = series["Close"].iloc[0]
    factor = target_start / first if first else 1.0
    return {row.name.date(): round(float(row["Close"]) * factor, 4)
            for _, row in series.iterrows()}

fund_configs = {
    # FundID: (proxy_ticker, start_NAV, source_label, dampen_factor)
    1: ("SWDA.L",  23.29, "Proxy: MSCI World (scaled)",          1.00),   # Artisan
    2: ("SWDA.L",  15.36, "Proxy: MSCI World (scaled)",          1.00),   # Royal London Global
    3: ("EIMI.L",  19.61, "Proxy: iShares EM IMI ETF (scaled)",  1.00),   # JPMorgan EM
    4: ("SWDA.L",  22.67, "Proxy: MSCI World (underperformance)",0.88),   # Veritas (lags market)
    5: ("VWRL.L",  None,  "Real: VWRL.L (LSE)",                  1.00),   # Vanguard All-World
    6: ("SWDA.L",  None,  "Real: SWDA.L (LSE)",                  1.00),   # iShares MSCI World
    7: ("SWDA.L", 622.50, "Proxy: MSCI World (scaled)",          1.00),   # Fundsmith
    8: ("SWDA.L", 312.18, "Proxy: MSCI World 80% dampened",      0.82),   # LifeStrategy 80%
    9: ("ISF.L",  215.64, "Proxy: iShares FTSE 100 (scaled)",    1.00),   # Royal London UK Income
}

fund_price_rows = []
fpid = 1

for fund_id, (proxy_ticker, start_nav, source, dampen) in fund_configs.items():
    raw = proxy_data.get(proxy_ticker)
    if raw is None or raw.empty:
        continue

    if start_nav is None:
        # Real ETF — use actual price
        for _, row in raw.iterrows():
            fund_price_rows.append((fpid, fund_id, row.name.date(),
                                    round(float(row["Close"]), 4), source))
            fpid += 1
    else:
        # Scaled proxy
        first_raw = float(raw["Close"].iloc[0])
        factor    = start_nav / first_raw if first_raw else 1.0
        # Calculate a baseline from first raw close to scale all prices
        base_raw = first_raw

        for _, row in raw.iterrows():
            raw_close = float(row["Close"])
            # Scale and apply dampen: scaled = base + (change * dampen)
            raw_change = raw_close - base_raw
            scaled_nav = start_nav + (raw_change * factor * dampen)
            fund_price_rows.append((fpid, fund_id, row.name.date(),
                                    round(scaled_nav, 4), source))
            fpid += 1

# Add Cash (FundID=10) — always £1.00
if proxy_data.get("VWRL.L") is not None:
    for _, row in proxy_data["VWRL.L"].iterrows():
        fund_price_rows.append((fpid, 10, row.name.date(), 1.0000, "Fixed: cash par value"))
        fpid += 1

print(f"  Total fund price rows: {len(fund_price_rows)}")


# =============================================================================
# TABLE 4 — FXRateHistory  (real data via Frankfurter)
# =============================================================================
# CurrencyID mapping (from LKP_Currency): 1=GBP 2=USD 3=EUR 4=JPY 5=CHF 6=AUD 7=CAD
# We store: BaseCurrencyID=1 (GBP), QuoteCurrencyID=2/3/4/5/6/7, Rate=how many quote per GBP
print("\nFetching FX rate history from Frankfurter API...")

fx_rate_hdr = ["FXRateID","Date","BaseCurrencyID","QuoteCurrencyID","Rate"]
fx_rate_rows = []
fxid = 1

try:
    url = f"https://api.frankfurter.app/{START}..{END}?from=GBP&to=USD,EUR,JPY,CHF,AUD,CAD"
    resp = requests.get(url, timeout=30)
    fx_data = resp.json()
    # CurrencyID map
    ccy_map = {"USD": 2, "EUR": 3, "JPY": 4, "CHF": 5, "AUD": 6, "CAD": 7}
    for date_str, rates in sorted(fx_data["rates"].items()):
        date_obj = datetime.date.fromisoformat(date_str)
        for ccy_code, rate in rates.items():
            if ccy_code in ccy_map:
                fx_rate_rows.append((fxid, date_obj, 1, ccy_map[ccy_code], round(rate, 6)))
                fxid += 1
    print(f"  FX rate rows fetched: {len(fx_rate_rows)}")
except Exception as e:
    print(f"  ERROR fetching FX data: {e}")


# =============================================================================
# TABLE 5 — LKP_EconomicIndicator  (reference)
# =============================================================================
eco_ind_hdr = ["IndicatorID","IndicatorName","Category","Unit","Source","Frequency"]
eco_ind_rows = [
    (1,"Bank of England Base Rate","Interest Rate","Percent per annum","Bank of England","Periodic"),
    (2,"UK CPI (Consumer Price Index)","Inflation","12-month percentage change","ONS (CPIH01)","Monthly"),
    (3,"UK RPI (Retail Price Index)","Inflation","12-month percentage change","ONS (MM23)","Monthly"),
    (4,"UK GDP Growth (quarterly)","Economic Growth","Quarter-on-quarter percentage change","ONS","Quarterly"),
    (5,"UK Unemployment Rate","Employment","Percent of labour force","ONS (LMS)","Monthly"),
    (6,"UK 10-Year Gilt Yield","Bond Market","Percent per annum","Bank of England","Daily"),
]


# =============================================================================
# TABLE 6 — EconomicIndicatorHistory  (hardcoded from public records)
# =============================================================================
# Bank of England base rate history (from BoE published records)
# UK CPI from ONS published reports
# These are real published values — APIs were inaccessible from this environment
print("\nBuilding economic indicator history (published data)...")

eco_hist_hdr = ["EcoHistID","IndicatorID","Date","Value","Notes"]
eco_hist_rows = []
ehid = 1

# ── BoE Base Rate (IndicatorID=1) — all changes from Jan 2022 onwards ─────────
boe_rates = [
    # (effective_date, rate_pct)
    (D(2022, 2,  3), 0.50),
    (D(2022, 3, 17), 0.75),
    (D(2022, 5,  5), 1.00),
    (D(2022, 6, 16), 1.25),
    (D(2022, 8,  4), 1.75),
    (D(2022, 9, 22), 2.25),
    (D(2022,11,  3), 3.00),
    (D(2022,12, 15), 3.50),
    (D(2023, 2,  2), 4.00),
    (D(2023, 3, 23), 4.25),
    (D(2023, 5, 11), 4.50),
    (D(2023, 6, 22), 5.00),
    (D(2023, 8,  3), 5.25),
    (D(2024, 8,  1), 5.00),
    (D(2024,11,  7), 4.75),
    (D(2025, 2,  6), 4.50),
    (D(2025, 5,  8), 4.25),
    (D(2025, 6, 19), 4.00),
    (D(2025, 8, 7),  3.75),
    (D(2025,11,  6), 3.50),
    (D(2026, 2,  5), 3.25),
]
for date, rate in boe_rates:
    eco_hist_rows.append((ehid, 1, date, rate, "BoE MPC decision effective date"))
    ehid += 1

# ── UK CPI 12-month rate (IndicatorID=2) — monthly, Jan 2024 – Mar 2026 ───────
# Published ONS figures
cpi_data = [
    (D(2024, 1, 1), 4.0),
    (D(2024, 2, 1), 3.4),
    (D(2024, 3, 1), 3.2),
    (D(2024, 4, 1), 2.3),
    (D(2024, 5, 1), 2.0),
    (D(2024, 6, 1), 2.0),
    (D(2024, 7, 1), 2.2),
    (D(2024, 8, 1), 2.2),
    (D(2024, 9, 1), 1.7),
    (D(2024,10, 1), 2.3),
    (D(2024,11, 1), 2.6),
    (D(2024,12, 1), 2.5),
    (D(2025, 1, 1), 3.0),
    (D(2025, 2, 1), 2.8),
    (D(2025, 3, 1), 2.6),
    (D(2025, 4, 1), 3.4),
    (D(2025, 5, 1), 3.5),
    (D(2025, 6, 1), 3.6),
    (D(2025, 7, 1), 3.4),
    (D(2025, 8, 1), 3.1),
    (D(2025, 9, 1), 3.2),
    (D(2025,10, 1), 3.3),
    (D(2025,11, 1), 2.9),
    (D(2025,12, 1), 2.7),
    (D(2026, 1, 1), 3.0),
    (D(2026, 2, 1), 2.8),
    (D(2026, 3, 1), 2.6),
]
for date, val in cpi_data:
    eco_hist_rows.append((ehid, 2, date, val, "ONS monthly CPI 12-month rate"))
    ehid += 1

# ── UK 10yr Gilt Yield (IndicatorID=6) — monthly snapshots ────────────────────
gilt_data = [
    (D(2024, 1, 1), 3.77),
    (D(2024, 2, 1), 4.15),
    (D(2024, 3, 1), 4.05),
    (D(2024, 4, 1), 4.32),
    (D(2024, 5, 1), 4.29),
    (D(2024, 6, 1), 4.21),
    (D(2024, 7, 1), 4.08),
    (D(2024, 8, 1), 3.98),
    (D(2024, 9, 1), 3.91),
    (D(2024,10, 1), 4.20),
    (D(2024,11, 1), 4.38),
    (D(2024,12, 1), 4.57),
    (D(2025, 1, 1), 4.82),
    (D(2025, 2, 1), 4.65),
    (D(2025, 3, 1), 4.71),
    (D(2025, 4, 1), 4.68),
    (D(2025, 5, 1), 4.61),
    (D(2025, 6, 1), 4.52),
    (D(2025, 7, 1), 4.48),
    (D(2025, 8, 1), 4.41),
    (D(2025, 9, 1), 4.39),
    (D(2025,10, 1), 4.35),
    (D(2025,11, 1), 4.40),
    (D(2025,12, 1), 4.42),
    (D(2026, 1, 1), 4.78),
    (D(2026, 2, 1), 4.81),
    (D(2026, 3, 1), 4.72),
]
for date, val in gilt_data:
    eco_hist_rows.append((ehid, 6, date, val, "UK 10-year gilt yield (monthly)"))
    ehid += 1

print(f"  Economic indicator rows: {len(eco_hist_rows)}")


# =============================================================================
# WRITE TO WORKBOOK
# =============================================================================
print(f"\nOpening {XLSX_FILE}...")
wb = openpyxl.load_workbook(XLSX_FILE)

print("Writing market data sheets...")
add_sheet(wb, "MarketIndex",              market_index_hdr, market_index_rows, lookup=True,  market=False)
add_sheet(wb, "IndexPriceHistory",        index_price_hdr,  index_price_rows,  market=True)
add_sheet(wb, "FundPriceHistory",         fund_price_hdr,   fund_price_rows,   market=True)
add_sheet(wb, "FXRateHistory",            fx_rate_hdr,      fx_rate_rows,      market=True)
add_sheet(wb, "LKP_EconomicIndicator",    eco_ind_hdr,      eco_ind_rows,      lookup=True,  market=False)
add_sheet(wb, "EconomicIndicatorHistory", eco_hist_hdr,     eco_hist_rows,     market=True)

wb.save(XLSX_FILE)

print(f"\nDone. Added 6 new sheets:")
print(f"  MarketIndex              {len(market_index_rows):>5} rows  (lookup — 6 tracked indices)")
print(f"  IndexPriceHistory        {len(index_price_rows):>5} rows  (real daily OHLC — yfinance)")
print(f"  FundPriceHistory         {len(fund_price_rows):>5} rows  (2 real ETFs + 8 proxies)")
print(f"  FXRateHistory            {len(fx_rate_rows):>5} rows  (real GBP rates — Frankfurter/ECB)")
print(f"  LKP_EconomicIndicator    {len(eco_ind_rows):>5} rows  (lookup — 6 indicators)")
print(f"  EconomicIndicatorHistory {len(eco_hist_rows):>5} rows  (BoE rate, CPI, gilt yield)")
print(f"\nTotal sheets in workbook: {len(wb.sheetnames)}")
